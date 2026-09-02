"""
Report views: PDF / Excel exports.

- GET /api/v1/reports/channel/<id>/pdf/   → medical case (channel) report
- GET /api/v1/reports/audit/excel/        → audit logs workbook (xlsx)
- GET /api/v1/reports/monthly/pdf/        → monthly hospital performance report

All exports are permission-checked and audited.
"""
import io
from datetime import datetime, timedelta

import matplotlib
matplotlib.use('Agg')  # headless
import matplotlib.pyplot as plt

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from rest_framework import permissions
from rest_framework.views import APIView
from rest_framework.response import Response

import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
)

from apps.accounts.models import User
from apps.channels.models import Channel, ChannelMessage
from apps.patients.models import Patient, MedicalRecord, MedicalFile
from apps.audit.models import AuditLog
from apps.audit.utils import log_security_event
from apps.security.permissions import IsAdmin, IsAuditor

from .arabic import ar, register_fonts, FONT_NORMAL, FONT_BOLD

BRAND = colors.HexColor('#2563EB')
BRAND_DARK = colors.HexColor('#1E40AF')
LIGHT_ROW = colors.HexColor('#EFF6FF')
GRID = colors.HexColor('#BFDBFE')


def _mpl_ar(text: str) -> str:
    """Arabic shaping for matplotlib labels."""
    return get_display(arabic_reshaper.reshape(str(text)))


class BaseReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def _audited_response(self, request, filename: str, content: bytes,
                          content_type: str, report_kind: str):
        log_security_event(
            user=request.user,
            event_type='VULN_SCAN_EXECUTED',  # keep enum stable; kind in details
            request=request,
            details={'report': report_kind, 'file': filename},
        )
        resp = HttpResponse(content, content_type=content_type)
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp


# ============================================================
# 1) Channel (medical case) PDF report
# ============================================================

class ChannelReportPDFView(BaseReportView):
    """Formatted PDF report for one medical case channel."""

    def get(self, request, channel_id: str):
        try:
            channel = Channel.objects.select_related('patient', 'owner').get(
                id=channel_id
            )
        except Channel.DoesNotExist:
            return Response({'detail': 'القناة غير موجودة'}, status=404)
        if not channel.can_view(request.user):
            return Response({'detail': 'غير مصرح لك بتقرير هذه القناة'}, status=403)

        register_fonts()
        patient: Patient = channel.patient
        memberships = channel.memberships.filter(is_active=True).select_related('user')
        records = MedicalRecord.objects.filter(channel=channel).order_by('-created_at')
        messages_count = ChannelMessage.objects.filter(channel=channel).count()
        files = MedicalFile.objects.filter(channel=channel)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
            leftMargin=15 * mm, rightMargin=15 * mm,
            title='SecureMed Channel Report',
        )

        title_style = ParagraphStyle(
            'title', fontName=FONT_BOLD, fontSize=17, alignment=2,
            textColor=BRAND_DARK, spaceAfter=10,
        )
        sub_style = ParagraphStyle(
            'sub', fontName=FONT_NORMAL, fontSize=10, alignment=2,
            textColor=colors.HexColor('#6B7280'),
        )
        head_style = ParagraphStyle(
            'head', fontName=FONT_BOLD, fontSize=12, alignment=2,
            textColor=BRAND_DARK, spaceBefore=10, spaceAfter=4,
        )
        cell = ParagraphStyle(
            'cell', fontName=FONT_NORMAL, fontSize=9, leading=12, alignment=2,
        )

        story = [
            Paragraph(ar('تقرير الحالة الطبية'), title_style),
            Paragraph(ar(f'SecureMed — منصة الرعاية الصحية الآمنة | {timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}'), sub_style),
            Spacer(1, 6 * mm),
        ]

        def info_table(rows, widths=(0.28, 0.72)):
            data = [[Paragraph(ar(k), ParagraphStyle('k', parent=cell, fontName=FONT_BOLD)), Paragraph(ar(v), cell)] for k, v in rows]
            t = Table(data, colWidths=[160 * mm * widths[0], 160 * mm * widths[1]])
            t.setStyle(TableStyle([
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, LIGHT_ROW]),
                ('GRID', (0, 0), (-1, -1), 0.5, GRID),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            return t

        def list_table(headers, rows, widths_mm):
            head = [Paragraph(ar(h), ParagraphStyle('h', parent=cell, fontName=FONT_BOLD, textColor=colors.white)) for h in headers]
            body = [[Paragraph(ar(v), cell) for v in r] for r in rows]
            t = Table([head] + body, colWidths=[w * mm for w in widths_mm], repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BRAND),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
                ('GRID', (0, 0), (-1, -1), 0.5, GRID),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3.5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
            ]))
            return t

        story.append(Paragraph(ar('معلومات القناة'), head_style))
        story.append(info_table([
            ('اسم القناة', channel.name),
            ('النوع', channel.get_channel_type_display()),
            ('الحالة', channel.get_status_display()),
            ('الأولوية', channel.get_priority_display()),
            ('مالك القناة', f'{channel.owner.full_name} ({channel.owner.email})'),
            ('تاريخ الإنشاء', timezone.localtime(channel.created_at).strftime('%Y-%m-%d')),
            ('عدد الرسائل', str(messages_count)),
        ]))

        story.append(Paragraph(ar('بيانات المريض'), head_style))
        age = ''
        if patient.date_of_birth:
            today = timezone.now().date()
            age = str(today.year - patient.date_of_birth.year - (
                (today.month, today.day) < (patient.date_of_birth.month, patient.date_of_birth.day)
            ))
        story.append(info_table([
            ('الاسم', patient.full_name),
            ('رقم الهوية', patient.national_id or '—'),
            ('تاريخ الميلاد', str(patient.date_of_birth or '—')),
            ('العمر', age),
            ('الجنس', patient.get_gender_display()),
            ('فصيلة الدم', patient.blood_type or '—'),
            ('الحساسية', patient.allergies or 'لا يوجد'),
        ]))

        story.append(Paragraph(ar(f'أعضاء الفريق ({memberships.count()})'), head_style))
        story.append(list_table(
            ['الاسم', 'الدور', 'تاريخ المنح'],
            [[m.user.full_name, m.get_role_display(),
              timezone.localtime(m.created_at).strftime('%Y-%m-%d')] for m in memberships],
            [60, 55, 45],
        ))

        story.append(Paragraph(ar(f'السجلات الطبية ({records.count()})'), head_style))
        story.append(list_table(
            ['النوع', 'العنوان', 'بواسطة', 'التاريخ', 'حرج؟'],
            [[r.get_record_type_display(), r.title,
              (r.created_by.full_name if r.created_by else '—'),
              timezone.localtime(r.created_at).strftime('%Y-%m-%d'),
              'نعم' if r.is_critical else 'لا'] for r in records[:30]],
            [35, 60, 35, 25, 15],
        ))

        if files.exists():
            story.append(Paragraph(ar(f'الملفات الطبية ({files.count()})'), head_style))
            story.append(list_table(
                ['العنوان', 'النوع', 'حجم (KB)'],
                [[f.title, f.get_file_type_display(), str(round(f.file_size / 1024))] for f in files[:20]],
                [70, 50, 30],
            ))

        footer = ParagraphStyle('f', parent=sub_style, alignment=1, spaceBefore=10)
        story.append(Paragraph(ar(f'تم التوليد بواسطة {request.user.full_name} — وثيقة سريرية محمية (HIPAA)'), footer))

        doc.build(story)
        log_security_event(
            user=request.user,
            event_type='PATIENT_DATA_ACCESSED',
            request=request,
            details={'report': 'channel_pdf', 'channel_id': str(channel.id)},
        )
        filename = f"SecureMed_Channel_{channel.name[:30].replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.pdf"
        return self._audited_response(
            request, filename, buf.getvalue(), 'application/pdf', 'channel_pdf'
        )


# ============================================================
# 2) Audit logs → Excel (xlsx)
# ============================================================

class AuditExcelView(BaseReportView):
    """Export audit logs to a styled xlsx workbook (admin/auditor)."""

    permission_classes = [permissions.IsAuthenticated, (IsAdmin | IsAuditor)]

    def get(self, request):
        qs = AuditLog.objects.select_related('user').order_by('-timestamp')

        severity = request.query_params.get('severity')
        event_type = request.query_params.get('event_type')
        after = request.query_params.get('timestamp_after')
        before = request.query_params.get('timestamp_before')
        if severity:
            qs = qs.filter(severity=severity)
        if event_type:
            qs = qs.filter(event_type=event_type)
        if after:
            qs = qs.filter(timestamp__gte=after)
        if before:
            qs = qs.filter(timestamp__lte=before)

        wb = Workbook()
        ws = wb.active
        ws.title = 'سجل التدقيق'
        ws.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='2563EB')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        headers = ['الوقت', 'المستخدم', 'نوع الحدث', 'الخطورة', 'عنوان IP', 'المسار', 'الطريقة']
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')

        for log in qs[:5000]:
            ws.append([
                timezone.localtime(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
                log.user.full_name if log.user else '—',
                log.get_event_type_display(),
                log.get_severity_display(),
                log.ip_address or '—',
                log.path or '—',
                log.method or '—',
            ])

        widths = [20, 24, 26, 12, 16, 28, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'

        # Summary sheet
        ws2 = wb.create_sheet('ملخص')
        ws2.rightToLeft = True
        ws2.append(['الإحصائية', 'القيمة'])
        ws2.append(['إجمالي السجلات المصدرة', qs.count()])
        for label, count in AuditLog.objects.values('severity').annotate(n=Count('id')).values_list('severity', 'n'):
            ws2.append([f'خطورة: {label}', count])
        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 14
        for col in ('A1', 'B1'):
            ws2[col].fill = header_fill
            ws2[col].font = header_font

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"SecureMed_Audit_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return self._audited_response(
            request, filename, buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'audit_excel',
        )


# ============================================================
# 3) Monthly performance report (PDF with charts)
# ============================================================

def _chart_bytes(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format='PNG', dpi=140, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf


class MonthlyReportPDFView(BaseReportView):
    """Monthly hospital report with matplotlib charts (admin/auditor)."""

    permission_classes = [permissions.IsAuthenticated, (IsAdmin | IsAuditor)]

    def get(self, request):
        from .monthly import build_monthly_report

        month_str = request.query_params.get('month')  # YYYY-MM
        try:
            pdf_bytes, filename, _start = build_monthly_report(
                month_str, generated_by=request.user.full_name
            )
        except ValueError:
            return Response({'detail': 'صيغة الشهر غير صحيحة (YYYY-MM)'}, status=400)

        return self._audited_response(
            request, filename, pdf_bytes, 'application/pdf', 'monthly_pdf'
        )


class MonthlyReportEmailView(BaseReportView):
    """
    POST /api/v1/reports/monthly/email/?month=YYYY-MM
    Generate the monthly PDF and email it (with attachment) to
    admins + auditors. The same operation runs on schedule via the
    management command `send_scheduled_reports --type monthly`.
    """

    permission_classes = [permissions.IsAuthenticated, (IsAdmin | IsAuditor)]

    def post(self, request):
        from .monthly import (
            build_monthly_report, monthly_report_recipients, send_report_email,
        )

        month_str = request.data.get('month') or request.query_params.get('month')
        try:
            pdf_bytes, filename, start = build_monthly_report(
                month_str, generated_by=request.user.full_name
            )
        except ValueError:
            return Response({'detail': 'صيغة الشهر غير صحيحة (YYYY-MM)'}, status=400)

        recipients = monthly_report_recipients()
        sent_to, failed = [], []
        for user in recipients:
            ok = send_report_email(
                user.email,
                start.strftime('%Y-%m'),
                filename,
                pdf_bytes,
            )
            (sent_to if ok else failed).append(user.email)

        log_security_event(
            user=request.user,
            event_type='MONTHLY_REPORT_EMAILED',
            request=request,
            details={
                'month': start.strftime('%Y-%m'),
                'recipients': sent_to,
                'failed': failed,
            },
        )
        return Response({
            'detail': f'تم إرسال التقرير إلى {len(sent_to)} مستلم بنجاح',
            'sent_to': sent_to,
            'failed': failed,
            'filename': filename,
        })


# ============================================================
# 4) Unified Report Download & List Views
# ============================================================

class ReportListView(APIView):
    """List all available system reports."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        reports = [
            {
                'id': 'monthly_summary',
                'title': 'التقرير الشهري الشامل',
                'description': 'إحصائيات شاملة للمرضى والقنوات والإيرادات للشهر المحدد',
                'formats': ['pdf', 'excel'],
            },
            {
                'id': 'patient_report',
                'title': 'تقرير المرضى',
                'description': 'قائمة شاملة بالمرضى وحالاتهم وسجلاتهم الطبية',
                'formats': ['pdf', 'excel'],
            },
            {
                'id': 'security_report',
                'title': 'تقرير الأمان',
                'description': 'سجل الأحداث الأمنية والثغرات المكتشفة وتوصيات الأمان',
                'formats': ['pdf', 'excel'],
            },
            {
                'id': 'appointments_report',
                'title': 'تقرير المواعيد',
                'description': 'إحصائيات المواعيد: معدلات الحضور، الإلغاء، وأوقات الذروة',
                'formats': ['pdf', 'excel'],
            },
            {
                'id': 'audit_report',
                'title': 'تقرير سجلات التدقيق',
                'description': 'سجل مفصّل لجميع العمليات والأحداث الأمنية في النظام',
                'formats': ['pdf', 'excel'],
            },
            {
                'id': 'channels_report',
                'title': 'تقرير القنوات / الحالات',
                'description': 'تحليل القنوات الطبية النشطة والمغلقة وتوزيعها',
                'formats': ['pdf', 'excel'],
            },
        ]
        return Response({'results': reports})


class UnifiedReportDownloadView(BaseReportView):
    """
    Unified report generation view for all reports.
    GET /api/v1/reports/<report_id>/?format=pdf|excel&start_date=...&end_date=...
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, report_id: str):
        format_type = request.query_params.get('format', 'pdf').lower()
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if report_id == 'monthly_summary':
            if format_type == 'excel':
                return self._export_monthly_excel(request, start_date, end_date)
            from .monthly import build_monthly_report
            month_str = start_date[:7] if start_date else timezone.now().strftime('%Y-%m')
            try:
                pdf_bytes, filename, _ = build_monthly_report(
                    month_str, generated_by=request.user.full_name
                )
                return self._audited_response(request, filename, pdf_bytes, 'application/pdf', 'monthly_summary')
            except Exception:
                return self._export_monthly_excel(request, start_date, end_date)

        elif report_id == 'audit_report':
            if format_type == 'excel':
                return AuditExcelView().get(request)
            return self._export_audit_pdf(request, start_date, end_date)

        elif report_id == 'patient_report':
            if format_type == 'excel':
                return self._export_patients_excel(request, start_date, end_date)
            return self._export_patients_pdf(request, start_date, end_date)

        elif report_id == 'appointments_report':
            if format_type == 'excel':
                return self._export_appointments_excel(request, start_date, end_date)
            return self._export_appointments_pdf(request, start_date, end_date)

        elif report_id == 'channels_report':
            if format_type == 'excel':
                return self._export_channels_excel(request, start_date, end_date)
            return self._export_channels_pdf(request, start_date, end_date)

        elif report_id == 'security_report':
            if format_type == 'excel':
                return self._export_security_excel(request, start_date, end_date)
            return self._export_security_pdf(request, start_date, end_date)

        return Response({'detail': f'التقرير غير معروف: {report_id}'}, status=404)

    def _export_patients_excel(self, request, start_date, end_date):
        qs = Patient.objects.all().order_by('-created_at')
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)

        wb = Workbook()
        ws = wb.active
        ws.title = 'قائمة المرضى'
        ws.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='059669')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        headers = ['اسم المريض', 'رقم الهوية', 'الجنس', 'فصيلة الدم', 'تاريخ الميلاد', 'الحوض الصحي', 'تاريخ التسجيل']
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')

        for p in qs[:2000]:
            ws.append([
                p.full_name or '—',
                p.national_id or '—',
                p.get_gender_display(),
                p.blood_type or '—',
                str(p.date_of_birth or '—'),
                p.basin.name if p.basin else '—',
                timezone.localtime(p.created_at).strftime('%Y-%m-%d'),
            ])

        for i, w in enumerate([25, 18, 10, 12, 14, 20, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"Patients_Report_{timezone.now().strftime('%Y%m%d')}.xlsx"
        return self._audited_response(
            request, filename, buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'patient_report_excel',
        )

    def _export_patients_pdf(self, request, start_date, end_date):
        register_fonts()
        qs = Patient.objects.all().order_by('-created_at')[:50]

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm)
        story = []

        title_style = ParagraphStyle('t', fontName=FONT_BOLD, fontSize=16, leading=20, alignment=1, textColor=BRAND)
        cell = ParagraphStyle('c', fontName=FONT_NORMAL, fontSize=8, leading=10, alignment=1)

        story.append(Paragraph(ar('تقرير سجلات المرضى المعتمد'), title_style))
        story.append(Spacer(1, 10 * mm))

        headers = ['تاريخ التسجيل', 'فصيلة الدم', 'الجنس', 'رقم الهوية', 'اسم المريض']
        rows = []
        for p in qs:
            rows.append([
                timezone.localtime(p.created_at).strftime('%Y-%m-%d'),
                p.blood_type or '—',
                p.get_gender_display(),
                p.national_id or '—',
                p.full_name or '—'
            ])

        head = [Paragraph(ar(h), ParagraphStyle('h', parent=cell, fontName=FONT_BOLD, textColor=colors.white)) for h in headers]
        body = [[Paragraph(ar(v), cell) for v in r] for r in rows]
        t = Table([head] + body, colWidths=[30 * mm, 25 * mm, 20 * mm, 35 * mm, 65 * mm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BRAND),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)

        doc.build(story)
        filename = f"Patients_Report_{timezone.now().strftime('%Y%m%d')}.pdf"
        return self._audited_response(request, filename, buf.getvalue(), 'application/pdf', 'patient_report_pdf')

    def _export_appointments_excel(self, request, start_date, end_date):
        from apps.appointments.models import Appointment
        qs = Appointment.objects.select_related('patient', 'doctor').all().order_by('-scheduled_at')
        if start_date:
            qs = qs.filter(scheduled_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(scheduled_at__date__lte=end_date)

        wb = Workbook()
        ws = wb.active
        ws.title = 'المواعيد الطبية'
        ws.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='7C3AED')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        headers = ['المريض', 'الطبيب', 'نوع الموعد', 'الحالة', 'الأولوية', 'الموعد المحدد', 'المكان']
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')

        for a in qs[:2000]:
            ws.append([
                a.patient.full_name if a.patient else '—',
                a.doctor.full_name if a.doctor else '—',
                a.get_appointment_type_display(),
                a.get_status_display(),
                a.get_priority_display(),
                timezone.localtime(a.scheduled_at).strftime('%Y-%m-%d %H:%M') if a.scheduled_at else '—',
                a.location or a.room_number or '—',
            ])

        for i, w in enumerate([22, 22, 16, 14, 12, 20, 18], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"Appointments_Report_{timezone.now().strftime('%Y%m%d')}.xlsx"
        return self._audited_response(
            request, filename, buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'appointments_excel',
        )

    def _export_appointments_pdf(self, request, start_date, end_date):
        from apps.appointments.models import Appointment
        register_fonts()
        qs = Appointment.objects.select_related('patient', 'doctor').all().order_by('-scheduled_at')[:50]

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm)
        story = []

        title_style = ParagraphStyle('t', fontName=FONT_BOLD, fontSize=16, leading=20, alignment=1, textColor=BRAND)
        cell = ParagraphStyle('c', fontName=FONT_NORMAL, fontSize=8, leading=10, alignment=1)

        story.append(Paragraph(ar('تقرير جدول المواعيد السريرية'), title_style))
        story.append(Spacer(1, 10 * mm))

        headers = ['الموعد', 'الأولوية', 'الحالة', 'النوع', 'الطبيب', 'المريض']
        rows = []
        for a in qs:
            rows.append([
                timezone.localtime(a.scheduled_at).strftime('%Y-%m-%d %H:%M') if a.scheduled_at else '—',
                a.get_priority_display(),
                a.get_status_display(),
                a.get_appointment_type_display(),
                a.doctor.full_name if a.doctor else '—',
                a.patient.full_name if a.patient else '—'
            ])

        head = [Paragraph(ar(h), ParagraphStyle('h', parent=cell, fontName=FONT_BOLD, textColor=colors.white)) for h in headers]
        body = [[Paragraph(ar(v), cell) for v in r] for r in rows]
        t = Table([head] + body, colWidths=[35 * mm, 20 * mm, 20 * mm, 25 * mm, 38 * mm, 38 * mm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)

        doc.build(story)
        filename = f"Appointments_Report_{timezone.now().strftime('%Y%m%d')}.pdf"
        return self._audited_response(request, filename, buf.getvalue(), 'application/pdf', 'appointments_pdf')

    def _export_channels_excel(self, request, start_date, end_date):
        qs = Channel.objects.select_related('patient', 'owner').all().order_by('-created_at')
        if start_date:
            qs = qs.filter(created_at__date__gte=start_date)
        if end_date:
            qs = qs.filter(created_at__date__lte=end_date)

        wb = Workbook()
        ws = wb.active
        ws.title = 'قنوات الحالات'
        ws.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='0891B2')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        headers = ['اسم القناة', 'المريض', 'مالك القناة', 'النوع', 'الحالة', 'الأولوية', 'تاريخ الإنشاء']
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')

        for ch in qs[:2000]:
            ws.append([
                ch.name,
                ch.patient.full_name if ch.patient else '—',
                ch.owner.full_name if ch.owner else '—',
                ch.get_channel_type_display(),
                ch.get_status_display(),
                ch.get_priority_display(),
                timezone.localtime(ch.created_at).strftime('%Y-%m-%d'),
            ])

        for i, w in enumerate([26, 22, 22, 16, 14, 12, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"Channels_Report_{timezone.now().strftime('%Y%m%d')}.xlsx"
        return self._audited_response(
            request, filename, buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'channels_excel',
        )

    def _export_channels_pdf(self, request, start_date, end_date):
        register_fonts()
        qs = Channel.objects.select_related('patient', 'owner').all().order_by('-created_at')[:50]

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm)
        story = []

        title_style = ParagraphStyle('t', fontName=FONT_BOLD, fontSize=16, leading=20, alignment=1, textColor=BRAND)
        cell = ParagraphStyle('c', fontName=FONT_NORMAL, fontSize=8, leading=10, alignment=1)

        story.append(Paragraph(ar('تقرير القنوات الطبية والحالات'), title_style))
        story.append(Spacer(1, 10 * mm))

        headers = ['التاريخ', 'الأولوية', 'الحالة', 'النوع', 'المريض', 'اسم القناة']
        rows = []
        for ch in qs:
            rows.append([
                timezone.localtime(ch.created_at).strftime('%Y-%m-%d'),
                ch.get_priority_display(),
                ch.get_status_display(),
                ch.get_channel_type_display(),
                ch.patient.full_name if ch.patient else '—',
                ch.name[:25]
            ])

        head = [Paragraph(ar(h), ParagraphStyle('h', parent=cell, fontName=FONT_BOLD, textColor=colors.white)) for h in headers]
        body = [[Paragraph(ar(v), cell) for v in r] for r in rows]
        t = Table([head] + body, colWidths=[25 * mm, 20 * mm, 20 * mm, 25 * mm, 38 * mm, 48 * mm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0891B2')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)

        doc.build(story)
        filename = f"Channels_Report_{timezone.now().strftime('%Y%m%d')}.pdf"
        return self._audited_response(request, filename, buf.getvalue(), 'application/pdf', 'channels_pdf')

    def _export_security_excel(self, request, start_date, end_date):
        qs = AuditLog.objects.filter(
            severity__in=['WARNING', 'CRITICAL']
        ).select_related('user').order_by('-timestamp')

        wb = Workbook()
        ws = wb.active
        ws.title = 'تنبيهات الأمان'
        ws.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='DC2626')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        headers = ['الوقت', 'المستخدم', 'الحدث الأمني', 'الخطورة', 'عنوان IP', 'بصمة الجهاز', 'المسار']
        ws.append(headers)
        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')

        for log in qs[:2000]:
            ws.append([
                timezone.localtime(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
                log.user.full_name if log.user else '—',
                log.get_event_type_display(),
                log.get_severity_display(),
                log.ip_address or '—',
                log.device_fingerprint[:16] if log.device_fingerprint else '—',
                log.path or '—',
            ])

        for i, w in enumerate([20, 22, 24, 12, 16, 20, 24], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"Security_Report_{timezone.now().strftime('%Y%m%d')}.xlsx"
        return self._audited_response(
            request, filename, buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'security_excel',
        )

    def _export_security_pdf(self, request, start_date, end_date):
        register_fonts()
        qs = AuditLog.objects.filter(severity__in=['WARNING', 'CRITICAL'])[:50]

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm)
        story = []

        title_style = ParagraphStyle('t', fontName=FONT_BOLD, fontSize=16, leading=20, alignment=1, textColor=colors.HexColor('#DC2626'))
        cell = ParagraphStyle('c', fontName=FONT_NORMAL, fontSize=8, leading=10, alignment=1)

        story.append(Paragraph(ar('تقرير الحوادث الأمنية والتهديدات'), title_style))
        story.append(Spacer(1, 10 * mm))

        headers = ['الوقت', 'المسار', 'عنوان IP', 'الخطورة', 'الحدث']
        rows = []
        for log in qs:
            rows.append([
                timezone.localtime(log.timestamp).strftime('%Y-%m-%d %H:%M'),
                log.path[:20] if log.path else '—',
                log.ip_address or '—',
                log.get_severity_display(),
                log.get_event_type_display()[:25]
            ])

        head = [Paragraph(ar(h), ParagraphStyle('h', parent=cell, fontName=FONT_BOLD, textColor=colors.white)) for h in headers]
        body = [[Paragraph(ar(v), cell) for v in r] for r in rows]
        t = Table([head] + body, colWidths=[35 * mm, 35 * mm, 30 * mm, 25 * mm, 50 * mm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)

        doc.build(story)
        filename = f"Security_Report_{timezone.now().strftime('%Y%m%d')}.pdf"
        return self._audited_response(request, filename, buf.getvalue(), 'application/pdf', 'security_pdf')

    def _export_audit_pdf(self, request, start_date, end_date):
        register_fonts()
        qs = AuditLog.objects.all().order_by('-timestamp')[:50]

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm)
        story = []

        title_style = ParagraphStyle('t', fontName=FONT_BOLD, fontSize=16, leading=20, alignment=1, textColor=BRAND)
        cell = ParagraphStyle('c', fontName=FONT_NORMAL, fontSize=8, leading=10, alignment=1)

        story.append(Paragraph(ar('تقرير سجلات التدقيق الأمني الشامل'), title_style))
        story.append(Spacer(1, 10 * mm))

        headers = ['الوقت', 'المسار', 'عنوان IP', 'الخطورة', 'المستخدم', 'الحدث']
        rows = []
        for log in qs:
            rows.append([
                timezone.localtime(log.timestamp).strftime('%m-%d %H:%M'),
                log.path[:15] if log.path else '—',
                log.ip_address or '—',
                log.get_severity_display(),
                (log.user.full_name[:15] if log.user else '—'),
                log.get_event_type_display()[:20]
            ])

        head = [Paragraph(ar(h), ParagraphStyle('h', parent=cell, fontName=FONT_BOLD, textColor=colors.white)) for h in headers]
        body = [[Paragraph(ar(v), cell) for v in r] for r in rows]
        t = Table([head] + body, colWidths=[25 * mm, 30 * mm, 25 * mm, 20 * mm, 35 * mm, 40 * mm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BRAND),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
            ('GRID', (0, 0), (-1, -1), 0.5, GRID),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(t)

        doc.build(story)
        filename = f"Audit_Report_{timezone.now().strftime('%Y%m%d')}.pdf"
        return self._audited_response(request, filename, buf.getvalue(), 'application/pdf', 'audit_pdf')

    def _export_monthly_excel(self, request, start_date, end_date):
        wb = Workbook()
        ws = wb.active
        ws.title = 'التقرير الشهري'
        ws.rightToLeft = True

        header_fill = PatternFill('solid', fgColor='1E40AF')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        headers = ['المؤشر السريري', 'القيمة الإحصائية']
        ws.append(headers)
        for col in (1, 2):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')

        from apps.appointments.models import Appointment
        total_patients = Patient.objects.count()
        total_channels = Channel.objects.count()
        active_channels = Channel.objects.filter(is_active=True).count()
        total_appointments = Appointment.objects.count()
        total_logs = AuditLog.objects.count()

        ws.append(['إجمالي المرضى المسجلين', total_patients])
        ws.append(['إجمالي القنوات السريرية', total_channels])
        ws.append(['القنوات النشطة حالياً', active_channels])
        ws.append(['إجمالي المواعيد المحجوزة', total_appointments])
        ws.append(['سجلات التدقيق الأمني المسجلة', total_logs])
        ws.append(['تاريخ توليد التقرير', timezone.now().strftime('%Y-%m-%d %H:%M')])
        ws.append(['المسؤول المنشئ', request.user.full_name])

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"Monthly_Summary_{timezone.now().strftime('%Y%m')}.xlsx"
        return self._audited_response(
            request, filename, buf.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'monthly_excel',
        )

